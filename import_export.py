"""
Расширенный модуль импорта/экспорта карточек
Поддержка: CSV, JSON, Quizlet, Google Sheets, Excel и т.д.
"""

import csv
import json
import re
from typing import List, Dict, Tuple, Optional
from abc import ABC, abstractmethod
import asyncio

class CardImporterBase(ABC):
    """Базовый класс для импортеров"""
    
    @abstractmethod
    async def import_cards(self, source: str) -> List[Tuple[str, str]]:
        """Импорт карточек из источника"""
        pass

class CSVImporter(CardImporterBase):
    """Импортер CSV файлов"""
    
    async def import_cards(self, file_path: str, 
                          question_col: int = 0, 
                          answer_col: int = 1,
                          skip_header: bool = True) -> List[Tuple[str, str]]:
        """
        Импорт из CSV файла
        
        Args:
            file_path: Путь к CSV файлу
            question_col: Индекс колонки с вопросами
            answer_col: Индекс колонки с ответами
            skip_header: Пропускать ли первую строку
        
        Returns:
            Список кортежей (вопрос, ответ)
        """
        cards = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                
                if skip_header:
                    next(reader)
                
                for row in reader:
                    if len(row) > max(question_col, answer_col):
                        question = row[question_col].strip()
                        answer = row[answer_col].strip()
                        
                        if question and answer:
                            cards.append((question, answer))
            
            print(f"✅ Импортировано из CSV: {len(cards)} карточек")
            return cards
        
        except Exception as e:
            print(f"❌ Ошибка при импорте CSV: {e}")
            return []

class JSONImporter(CardImporterBase):
    """Импортер JSON файлов"""
    
    async def import_cards(self, file_path: str) -> List[Tuple[str, str]]:
        """
        Импорт из JSON файла
        
        Ожидаемый формат:
        [{"question": "q1", "answer": "a1"}, ...]
        или
        {"cards": [{"question": "q1", "answer": "a1"}, ...]}
        """
        cards = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            
            # Если данные в ключе 'cards'
            if isinstance(data, dict) and 'cards' in data:
                data = data['cards']
            
            # Если это список словарей
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict):
                        question = item.get('question', '').strip()
                        answer = item.get('answer', '').strip()
                        
                        if question and answer:
                            cards.append((question, answer))
            
            print(f"✅ Импортировано из JSON: {len(cards)} карточек")
            return cards
        
        except Exception as e:
            print(f"❌ Ошибка при импорте JSON: {e}")
            return []

class TextImporter(CardImporterBase):
    """Импортер из простого текста"""
    
    async def import_cards(self, text: str, separator: str = '|') -> List[Tuple[str, str]]:
        """
        Импорт из текста
        
        Формат: каждая строка - одна карточка
        Вопрос|Ответ
        """
        cards = []
        
        for line in text.strip().split('\n'):
            if not line.strip():
                continue
            
            if separator in line:
                parts = line.split(separator)
                if len(parts) >= 2:
                    question = parts[0].strip()
                    answer = parts[1].strip()
                    
                    if question and answer:
                        cards.append((question, answer))
        
        print(f"✅ Импортировано из текста: {len(cards)} карточек")
        return cards

class QuizletImporter(CardImporterBase):
    """Импортер из Quizlet"""
    
    async def import_cards(self, quizlet_url: str) -> List[Tuple[str, str]]:
        """
        Импорт из Quizlet по URL
        Требует парсинга HTML страницы
        """
        try:
            import requests
            from bs4 import BeautifulSoup
            
            # Получение HTML
            response = requests.get(quizlet_url, timeout=10)
            
            if response.status_code != 200:
                print("❌ Не удалось подключиться к Quizlet")
                return []
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Поиск карточек (структура может отличаться)
            cards = []
            
            # Попытка найти скрипт с данными
            scripts = soup.find_all('script')
            for script in scripts:
                if 'studyable' in script.text:
                    # Парсинг JSON из скрипта
                    match = re.search(r'"terms":\s*(\[.*?\])', script.text)
                    if match:
                        try:
                            terms = json.loads(match.group(1))
                            for term in terms:
                                question = term.get('word', '').strip()
                                answer = term.get('definition', '').strip()
                                
                                if question and answer:
                                    cards.append((question, answer))
                        except:
                            pass
            
            if cards:
                print(f"✅ Импортировано из Quizlet: {len(cards)} карточек")
            else:
                print("⚠️ Карточки не найдены. Может потребоваться другой способ.")
            
            return cards
        
        except ImportError:
            print("⚠️ Требуются библиотеки: requests, beautifulsoup4")
            print("   Установите: pip install requests beautifulsoup4")
            return []
        except Exception as e:
            print(f"❌ Ошибка при импорте с Quizlet: {e}")
            return []

class ExcelImporter(CardImporterBase):
    """Импортер Excel файлов"""
    
    async def import_cards(self, file_path: str,
                          sheet_name: int = 0,
                          question_col: str = 'A',
                          answer_col: str = 'B') -> List[Tuple[str, str]]:
        """
        Импорт из Excel файла
        
        Args:
            file_path: Путь к файлу Excel
            sheet_name: Индекс листа
            question_col: Колонка с вопросами
            answer_col: Колонка с ответами
        """
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.worksheets[sheet_name]
            
            cards = []
            
            for row in worksheet.iter_rows(values_only=True):
                # Получение значений из колонок
                question_idx = ord(question_col.upper()) - ord('A')
                answer_idx = ord(answer_col.upper()) - ord('A')
                
                if len(row) > max(question_idx, answer_idx):
                    question = str(row[question_idx] or '').strip()
                    answer = str(row[answer_idx] or '').strip()
                    
                    if question and answer:
                        cards.append((question, answer))
            
            print(f"✅ Импортировано из Excel: {len(cards)} карточек")
            return cards
        
        except ImportError:
            print("⚠️ Требуется библиотека openpyxl")
            print("   Установите: pip install openpyxl")
            return []
        except Exception as e:
            print(f"❌ Ошибка при импорте Excel: {e}")
            return []

class GoogleSheetsImporter(CardImporterBase):
    """Импортер из Google Sheets"""
    
    async def import_cards(self, sheet_url: str) -> List[Tuple[str, str]]:
        """
        Импорт из Google Sheets
        
        URL должен быть доступным CSV экспортом:
        https://docs.google.com/spreadsheets/d/{ID}/export?format=csv
        """
        try:
            import requests
            
            # Преобразование URL в CSV экспорт если нужно
            if 'spreadsheets/d/' in sheet_url:
                sheet_id = sheet_url.split('/d/')[1].split('/')[0]
                csv_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
            else:
                csv_url = sheet_url
            
            response = requests.get(csv_url, timeout=10)
            
            if response.status_code != 200:
                print("❌ Не удалось получить доступ к Google Sheets")
                return []
            
            # Парсинг CSV
            lines = response.text.split('\n')
            cards = []
            
            for line in lines[1:]:  # Пропускаем заголовок
                if ',' in line:
                    parts = line.split(',', 1)
                    if len(parts) == 2:
                        question = parts[0].strip().strip('"')
                        answer = parts[1].strip().strip('"')
                        
                        if question and answer:
                            cards.append((question, answer))
            
            print(f"✅ Импортировано из Google Sheets: {len(cards)} карточек")
            return cards
        
        except Exception as e:
            print(f"❌ Ошибка при импорте с Google Sheets: {e}")
            return []


class CardExporter:
    """Экспортер карточек в различные форматы"""
    
    @staticmethod
    async def export_to_csv(cards: List[Dict], file_path: str) -> bool:
        """Экспорт в CSV"""
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as file:
                writer = csv.DictWriter(file, fieldnames=['question', 'answer', 'difficulty'])
                writer.writeheader()
                
                for card in cards:
                    writer.writerow({
                        'question': card.get('question', ''),
                        'answer': card.get('answer', ''),
                        'difficulty': card.get('difficulty', 1)
                    })
            
            print(f"✅ Экспортировано в CSV: {file_path}")
            return True
        
        except Exception as e:
            print(f"❌ Ошибка при экспорте в CSV: {e}")
            return False
    
    @staticmethod
    async def export_to_json(cards: List[Dict], file_path: str) -> bool:
        """Экспорт в JSON"""
        try:
            data = []
            
            for card in cards:
                data.append({
                    'question': card.get('question', ''),
                    'answer': card.get('answer', ''),
                    'difficulty': card.get('difficulty', 1)
                })
            
            with open(file_path, 'w', encoding='utf-8') as file:
                json.dump(data, file, ensure_ascii=False, indent=2)
            
            print(f"✅ Экспортировано в JSON: {file_path}")
            return True
        
        except Exception as e:
            print(f"❌ Ошибка при экспорте в JSON: {e}")
            return False
    
    @staticmethod
    async def export_to_excel(cards: List[Dict], file_path: str) -> bool:
        """Экспорт в Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            
            # Заголовок
            worksheet['A1'] = 'Вопрос'
            worksheet['B1'] = 'Ответ'
            worksheet['C1'] = 'Сложность'
            
            # Форматирование заголовка
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in ['A1', 'B1', 'C1']:
                worksheet[cell].fill = header_fill
                worksheet[cell].font = header_font
            
            # Данные
            for idx, card in enumerate(cards, start=2):
                worksheet[f'A{idx}'] = card.get('question', '')
                worksheet[f'B{idx}'] = card.get('answer', '')
                worksheet[f'C{idx}'] = card.get('difficulty', 1)
            
            # Регулировка ширины колонок
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 30
            worksheet.column_dimensions['C'].width = 12
            
            workbook.save(file_path)
            print(f"✅ Экспортировано в Excel: {file_path}")
            return True
        
        except ImportError:
            print("⚠️ Требуется библиотека openpyxl")
            return False
        except Exception as e:
            print(f"❌ Ошибка при экспорте в Excel: {e}")
            return False
    
    @staticmethod
    async def export_to_text(cards: List[Dict], file_path: str) -> bool:
        """Экспорт в текстовый файл"""
        try:
            with open(file_path, 'w', encoding='utf-8') as file:
                for i, card in enumerate(cards, 1):
                    file.write(f"{i}. {card.get('question', '')} | {card.get('answer', '')}\n")
            
            print(f"✅ Экспортировано в текст: {file_path}")
            return True
        
        except Exception as e:
            print(f"❌ Ошибка при экспорте в текст: {e}")
            return False
    
    @staticmethod
    async def export_for_quizlet(cards: List[Dict], file_path: str = None) -> str:
        """Генерирование текста для импорта в Quizlet"""
        text = ""
        
        for card in cards:
            question = card.get('question', '')
            answer = card.get('answer', '')
            text += f"{question}\t{answer}\n"
        
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as file:
                    file.write(text)
                print(f"✅ Экспортировано для Quizlet: {file_path}")
            except Exception as e:
                print(f"❌ Ошибка: {e}")
        
        return text


# Пример использования
if __name__ == "__main__":
    async def main():
        # Пример импорта из текста
        text = """What is Python | A programming language
What is AI | Artificial Intelligence
What is ML | Machine Learning"""
        
        importer = TextImporter()
        cards = await importer.import_cards(text)
        
        for q, a in cards:
            print(f"Q: {q} → A: {a}")
        
        # Пример экспорта
        print("\nЭкспорт...")
        cards_dict = [
            {'question': q, 'answer': a, 'difficulty': 1}
            for q, a in cards
        ]
        
        await CardExporter.export_to_json(cards_dict, 'exported_cards.json')
        await CardExporter.export_to_csv(cards_dict, 'exported_cards.csv')
    
    asyncio.run(main())
